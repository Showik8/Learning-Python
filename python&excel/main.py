import openpyxl

inv_file = openpyxl.load_workbook('in.xlsx')
product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}
product_list.cell(1,5).value="Total Value"


for product_row in range(2,product_list.max_row+1):
    supplier_name = product_list.cell(product_row,4,).value
    inventory_amount = product_list.cell(product_row,2).value
    price = product_list.cell(product_row,3).value
    inventory_price = product_list.cell(product_row,5)

    # Calculation of products 
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name] 
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
    
    # calculation value
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + inventory_amount * price
    else:
        total_value_per_supplier[supplier_name] = inventory_amount * price

    inventory_price.value = int(inventory_amount * price)

    print(inventory_price.value)

    inv_file.save("Made in python.xlsx")

        


